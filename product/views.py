import csv
import tempfile
import zipfile
from functools import reduce
from io import BytesIO
from tempfile import NamedTemporaryFile
from PIL import Image, ImageDraw, ImageFont

from django.db.models import Sum, Count, IntegerField, Q
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination

from rest_framework.response import Response

from product.filters import TypeFilter
from product.models import Product, Collectable, VideoGame, Accessory, Report, StateEnum, Sale
from product.serializer import ProductSerializer, CollectableSerializer, VideoGameSerializer, AccessorySerializer, \
    ReportSerializer
from datetime import datetime

from django.utils import timezone
from rest_framework.views import APIView
from rest_framework import filters, status
from django.http import JsonResponse, HttpResponse
from django.db.models.functions import Cast


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 12
    page_size_query_param = 'page_size'
    max_page_size = 12


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.filter(state=StateEnum.available)
    serializer_class = ProductSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = []

    def get_queryset(self):
        """
        Optionally restricts the returned purchases to a given user,
        by filtering against a `username` query parameter in the URL.
        """
        tags = self.request.query_params.get('tags')
        if not tags:
            return self.queryset

        tags = tags.split(",")
        return self.queryset.filter(tags__name__in=tags)

    def retrieve(self, request, *args, **kwargs):
        self.queryset = Product.objects.filter(state__in=[StateEnum.available, StateEnum.reserved])
        pk = kwargs.get("pk")
        instance = None

        if pk is not None:
            if pk.isdigit():
                instance = self.queryset.filter(id=pk).first()

            if instance is None:
                instance = self.queryset.filter(barcode=pk).first()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)


class CollectableViewSet(viewsets.ModelViewSet):
    queryset = Collectable.objects.filter()
    serializer_class = CollectableSerializer
    permission_classes = []


class VideoGameViewSet(viewsets.ModelViewSet):
    queryset = VideoGame.objects.filter()
    serializer_class = VideoGameSerializer
    permission_classes = []


class AccessoryViewSet(viewsets.ModelViewSet):
    queryset = Accessory.objects.filter()
    serializer_class = AccessorySerializer
    permission_classes = []


class DailySalesReport(APIView):

    def post(self, request):
        start_date_str = request.data.get('start_date')
        end_date_str = request.data.get('end_date')

        if start_date_str and end_date_str:
            try:
                start_date = timezone.make_aware(datetime.strptime(start_date_str, "%Y-%m-%d"))
                end_date = timezone.make_aware(datetime.strptime(end_date_str, "%Y-%m-%d"))
            except ValueError:
                return Response(
                    {"error": "Las fechas proporcionadas no tienen el formato correcto (YYYY-MM-DD)."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            queryset = Sale.objects.filter(purchase_date_time__range=(start_date, end_date))
            queryset = queryset.values('purchase_date_time__date').annotate(
                total_sales=Count('id'),
                gross_total=Cast(Sum('gross_total'), output_field=IntegerField()),
                net_total=Cast(Sum('net_total'), output_field=IntegerField()),
            ).order_by('purchase_date_time__date')

            data = list(queryset)  # Convertir el QuerySet en una lista de diccionarios
            return JsonResponse(data, safe=False)
        else:
            return Response(
                {"error": "Debes proporcionar las fechas de inicio y fin (start_date y end_date) en los parámetros de consulta."},
                status=status.HTTP_400_BAD_REQUEST
            )


class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    filter_backends = [filters.OrderingFilter]

    def get_queryset(self):
        queryset = super().get_queryset()

        date_param = self.request.query_params.get('date', None)
        if date_param:
            queryset = queryset.filter(date=date_param)

        return queryset.annotate(total_products=Sum('sale__products__sale_price'))


class GenerateExcelOfProducts(APIView):

    def get(self, request):
        new_file_object, _, console = self.get_products_and_create_temp_file(request)

        # Create an HttpResponse with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Get today's date
        today_date = datetime.now()

        # Format the date as a string in "DD-MM-YYYY" format
        formatted_date = today_date.strftime("%d-%m-%Y\ %H:%M")

        response['Content-Disposition'] = f'attachment; filename=productos\ {formatted_date}\ {console if console else ""}.xlsx'
        response.write(new_file_object)

        return response

    @staticmethod
    def get_products_and_create_temp_file(request):
        products = Product.objects.filter(state=StateEnum.available)

        console = request.query_params.get("console_title")
        product_type = request.query_params.get("type")
        used = request.query_params.get("used")
        product_name = request.query_params.get("name")

        if console:
            products = products.filter(
                Q(console__title=console) | Q(videogame__console=console) | Q(accessory__console=console)
            )
        if product_type:
            products = TypeFilter.get_products_by_type(product_type, products)
        if used and used.isnumeric():
            products = products.filter(used=int(used))
        if product_name:
            # List of fields to search in
            search_fields = ["videogame__title", "barcode", "console__title", "accessory__title", "collectable__title",
                             "description"]

            # Construct the OR condition using list comprehensions
            or_condition = reduce(lambda q, field: q | Q(**{f"{field}__icontains": product_name}), search_fields, Q())
            # Apply the filter to your queryset
            products = products.filter(or_condition)

        product_barcodes = Product.objects.values('barcode').distinct()
        products = products.filter(barcode__in=product_barcodes.values('barcode'))
        rows = [
            [str(product), product.sale_price, product.description, str(product.console_type),
             product.get_product_type()] for product in products
        ]
        wb = Workbook()
        # Add data to the Excel workbook (replace this with your actual data)
        sheet = wb.active
        sheet.title = "Productos"
        sheet['A1'] = "Nombre"
        sheet['B1'] = "Precio"
        sheet['C1'] = "Descripción"
        sheet['D1'] = "Consola"
        sheet['E1'] = "Tipo"
        # Make titles bold
        for cell in sheet['1:1']:
            cell.font = Font(bold=True)
        for row_num, row in enumerate(rows):
            sheet[f"A{row_num + 2}"] = row[0]
            sheet[f"B{row_num + 2}"] = f"₡{row[1]:,.2f}"
            sheet[f"C{row_num + 2}"] = row[2]
            sheet[f"D{row_num + 2}"] = row[3]
            sheet[f"E{row_num + 2}"] = row[4]

            # Set the column width based on content length
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        with NamedTemporaryFile() as tmp:
            tmp.close()  # with statement opened tmp, close it so wb.save can open it
            wb.save(tmp.name)
            with open(tmp.name, 'rb') as f:
                f.seek(0)  # probably not needed anymore
                new_file_object = f.read()
        return new_file_object, wb, console


class GenerateImageOfProducts(APIView):

    @staticmethod
    def chunk_list(input_list, chunk_size):
        """
        Split a list into chunks of a specified size.

        Parameters:
        - input_list: The list to be chunked.
        - chunk_size: The size of each chunk.

        Returns:
        A list of chunks, where each chunk is a sublist of the input_list.
        """
        return [input_list[i:i + chunk_size] for i in range(0, len(input_list), chunk_size)]

    def get(self, request):
        products, wb, console = GenerateExcelOfProducts.get_products_and_create_temp_file(request)

        # Create an HttpResponse with the Excel file
        response = HttpResponse(content_type='image/png')

        # Get today's date
        today_date = datetime.now()

        # Format the date as a string in "DD-MM-YYYY" format
        formatted_date = today_date.strftime("%d-%m-%Y\ %H:%M")

        # Create a temporary CSV file
        with tempfile.NamedTemporaryFile(mode='w+', delete=False, suffix='.csv') as temp_csv:
            csv_writer = csv.writer(temp_csv)

            # Iterate through all sheets in the workbook
            for sheet in wb.sheetnames:
                current_sheet = wb[sheet]
                for row in current_sheet.iter_rows(values_only=True):
                    csv_writer.writerow(row)

        with open(temp_csv.name, "r") as temp_csv:
            csv_reader = csv.reader(temp_csv)
            data = list(csv_reader)
            data_matrix = GenerateImageOfProducts.chunk_list(data, 60)

            # Set font properties
            font_size = 12
            font_path = './Arial.ttf'  # Replace with the path to a TrueType font file
            font = ImageFont.truetype(font_path, font_size)

            # Calculate image size based on the number of rows and columns
            row_height = 20
            column_width = 150
            image_width = len(data[0]) * column_width
            image_height = (len(data) + 1) * row_height  # +1 for header

            images = []
            for data in data_matrix:
                # Create a new image
                image = Image.new('RGB', (image_width, image_height), color='white')
                draw = ImageDraw.Draw(image)

                # Draw headers
                for i, header in enumerate(data[0]):
                    draw.text((i * column_width, 0), header, font=font, fill='black')

                # Draw data rows
                for row_num, row_data in enumerate(data[1:], start=1):
                    prev_x = 0
                    for col_num, cell_data in enumerate(row_data):
                        draw.text((col_num * column_width + prev_x, row_num * row_height), cell_data, font=font, fill='black')
                        prev_x = col_num * column_width + 30

                # Save the image as PNG in memory (BytesIO)
                image_bytes = BytesIO()
                image.save(image_bytes, format='PNG')
                image_bytes.seek(0)

                # Save BytesIO image as a temporary file
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
                    temp_file.write(image_bytes.read())

                images.append(image_bytes)

                if len(data) <= 60:
                    response['Content-Disposition'] = f'attachment; filename=productos\ {formatted_date}\ {console}.png'
                    response.write(image_bytes.getvalue())

                    return response

        def zip_files(files):
            outfile = BytesIO()
            with zipfile.ZipFile(outfile, 'w') as zf:
                for n, f in enumerate(files):
                    zf.writestr(f"productos_{formatted_date}_{console}_{n}.png", f.getvalue())
            return outfile.getvalue()

        zipped_file = zip_files(images)
        response = HttpResponse(zipped_file, content_type='application/octet-stream')
        response['Content-Disposition'] = 'attachment; filename=productos.zip'

        return response
